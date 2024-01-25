import json
from openpyxl import Workbook, load_workbook
from azure.core.credentials import AzureKeyCredential
from azure.ai.formrecognizer import DocumentAnalysisClient
import time
from tkinter import *
from tkinter.filedialog import askopenfilenames

credentials = json.load(open('credentials.json'))
API_KEY = credentials['API_KEY']
ENDPOINT = credentials['ENDPOINT']
document_analysis_client = DocumentAnalysisClient(ENDPOINT, AzureKeyCredential(API_KEY))


def select_files():
    files_selected = askopenfilenames(filetypes=(("PDF Documents", "*.pdf"),))
    return files_selected


def process_files(files_selected):
    wb = Workbook().active
    wb.save("test.xlsx")
    path = 'C:\\Projetos VSCODE\\PERDCOM PROJECT\\test.xlsx'
    data_sheet_inactive = load_workbook(filename=path)
    data_sheet = data_sheet_inactive.worksheets[0]
    data_sheet['A1'] = 'Nome'
    data_sheet['B1'] = 'CPF/CNPJ'
    file_line = 2
    for files in files_selected:
        with open(files, "rb") as file:
            poller = document_analysis_client.begin_analyze_document(model_id="prebuilt-document", document=file)
            result = poller.result()
            for kv_pair in result.key_value_pairs:
                if kv_pair.key.content == "Nome":
                    data_sheet[f'A{file_line}'] = kv_pair.value.content
                if kv_pair.key.content == 'CNPJ/CPF':
                    data_sheet[f'B{file_line}'] = kv_pair.value.content
            data_sheet_inactive.save('test.xlsx')
            time.sleep(65)
        file_line += 1
    print('terminou')


window = Tk()
window.geometry('800x800')

select_files_button = Button(text='Selecionar arquivos', command=select_files)
select_files_button.grid(column=0, row=0)

selected_files = LabelFrame

process_files_button = Button(text='Processar arquivos', command=process_files)
process_files_button.grid(column=0, row=1)

window.mainloop()

'''
with open('test.pdf', "rb") as i:
    poller = document_analysis_client.begin_analyze_document(model_id="prebuilt-document", document=i)
    result = poller.result()
    text = result.content
    lines = enumerate(text.split('\n'), start=1)
    data = {}
    wb = Workbook()
    ws = wb.active
    wb.save("test.xlsx")
    path = 'C:\\Projetos VSCODE\\PERDCOM PROJECT\\test.xlsx'
    data_sheet_inactive = load_workbook(filename=path)
    data_sheet = data_sheet_inactive.worksheets[0]
    for line, text_line in lines:
        print(f'index={line} | text of the line={text_line}')
        if line == 27:
            key = 'Nome'
            value = text_line
            data[key] = value
            data_sheet['A1'] = key
            data_sheet['A2'] = data[key]
        if line == 14:
            key = 'CPF'
            value = text_line
            data[key] = value
            data_sheet['B1'] = key
            data_sheet['B2'] = data[key]
    data_sheet_inactive.save('test.xlsx')
'''
